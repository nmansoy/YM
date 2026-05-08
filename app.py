import streamlit as st
import pandas as pd
import io
import json
import os
import openpyxl
from decimal import Decimal, ROUND_HALF_UP

# --- Excel Tarzı Kesin Yuvarlama Fonksiyonu ---
def excel_yuvarla(deger, basamak=2):
    format_str = '1.' + '0' * basamak if basamak > 0 else '1'
    return float(Decimal(str(deger)).quantize(Decimal(format_str), rounding=ROUND_HALF_UP))

# --- Veri Saklama ve Sıralama ---
DATA_FILE = "endeksler.json"
AYLAR_SOZLUK = {"Ocak": 1, "Şubat": 2, "Mart": 3, "Nisan": 4, "Mayıs": 5, "Haziran": 6, "Temmuz": 7, "Ağustos": 8, "Eylül": 9, "Ekim": 10, "Kasım": 11, "Aralık": 12}

def kronolojik_sirala(endeks_dict):
    def siralama_anahtari(item):
        try:
            ay_isim, yil = item[0].split()
            return int(yil) * 100 + AYLAR_SOZLUK.get(ay_isim, 0)
        except:
            return 0
    return dict(sorted(endeks_dict.items(), key=siralama_anahtari, reverse=True))

def endeksleri_yukle():
    if not os.path.exists(DATA_FILE):
        varsayilan = {"Mart 2026": 4747.63, "Şubat 2026": 4620.50}
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(varsayilan, f, ensure_ascii=False, indent=4)
        return kronolojik_sirala(varsayilan)
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return kronolojik_sirala(json.load(f))

def endeks_kaydet(donem_adi, deger):
    veriler = endeksleri_yukle()
    veriler[donem_adi] = float(deger)
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(kronolojik_sirala(veriler), f, ensure_ascii=False, indent=4)

AYLAR_LISTESI = list(AYLAR_SOZLUK.keys())
YILLAR_LISTESI = [str(y) for y in range(2022, 2035)]

# --- Sayfa Ayarları ve CSS İlaveleri ---
st.set_page_config(page_title="AYAP Maliyet Hesaplama", layout="wide")

st.markdown("""
    <style>
    .stRadio [role=radiogroup] {
        display: flex;
        justify-content: center;
        gap: 30px;
        background-color: #f0f2f6;
        padding: 20px;
        border-radius: 10px;
        border: 2px solid #ff4b4b;
    }
    .stRadio label [data-testid="stMarkdownContainer"] p {
        font-size: 24px !important;
        font-weight: 900 !important;
        color: #1f1f1f !important;
    }
    @media (prefers-color-scheme: dark) {
        .stRadio [role=radiogroup] { background-color: #262730; }
        .stRadio label [data-testid="stMarkdownContainer"] p { color: #ffffff !important; }
    }
    </style>
""", unsafe_allow_html=True)

# --- ANA EKRAN BAŞLIK VE UYARI ---
st.title("📊 AYAP Kalem Kalem Maliyet Hesaplama Aracı")

st.warning("""
🚨 **LÜTFEN DİKKAT - İŞLEME BAŞLAMADAN ÖNCE KONTROL EDİNİZ:**
1. Sol menüden **Mesh Üretimli** ya da **Mesh Üretimsiz** fiyat seçimini doğru yaptığınızdan emin olun.
2. Hesaplamanın doğru yapılabilmesi için sol menüdeki **Güncel Ay Endeksinin** çalışmanıza uygun ayı gösterdiğini teyit ediniz.
""", icon="⚠️")

# --- Sol Menü ---
st.sidebar.header("⚙️ Temel Parametreler")
st.sidebar.markdown("🔗 **[Yİ-ÜFE Endeksleri (hakedis.org)](https://www.hakedis.org/endeksler/yi-ufe-yurtici-uretici-fiyat-endeksi)**")
st.sidebar.markdown("---")

st.sidebar.subheader("🗺️ Pafta Üretim Tipi")
mesh_durumu = st.sidebar.radio("Mesh model üretimi yapılacak mı?", ("✅ Mesh Üretimli (Mevcut)", "❌ Mesh Üretimsiz (İptal Edilen)"))

base_endeks = 1210.60
base_bina = 76.82
base_pafta = 1247.814 if "Üretimli" in mesh_durumu else 1155.384

if "Üretimsiz" in mesh_durumu:
    st.sidebar.warning("⚠️ Mesh üretimi iptal edilmiş pafta baz fiyatı (1155.384 TL) kullanılır.")

st.sidebar.markdown("---")
endeks_verileri = endeksleri_yukle()
secenekler = list(endeks_verileri.keys())
secilen_donem = st.sidebar.selectbox("Endeks Seçin", options=secenekler, index=0)
guncel_endeks = endeks_verileri[secilen_donem]
st.sidebar.success(f"**{secilen_donem}:** {guncel_endeks}")

with st.sidebar.expander("➕ Yeni Endeks Ekle"):
    ay = st.selectbox("Ay", AYLAR_LISTESI)
    yil = st.selectbox("Yıl", YILLAR_LISTESI, index=4)
    deger = st.number_input("Değer", min_value=0.0, format="%.2f", step=10.0)
    if st.button("💾 Kaydet") and deger > 0:
        endeks_kaydet(f"{ay} {yil}", deger)
        st.rerun()

st.sidebar.markdown("---")
kar_orani = st.sidebar.number_input("Yüklenici Kârı (%)", value=20.0, step=1.0)
kdv_orani = st.sidebar.number_input("KDV Oranı (%)", value=20.0, step=1.0)

# --- Katsayı ve Fiyatlar ---
katsayi = guncel_endeks / base_endeks
guncel_bina = base_bina * katsayi
guncel_pafta = base_pafta * katsayi

st.markdown("### 📈 Güncel Baz Fiyatlar ve Çarpan")
m_col1, m_col2, m_col3 = st.columns(3)
m_col1.metric("Endeks Çarpanı", f"{katsayi:.5f}")
m_col2.metric("Güncel Bina B.F.", f"{guncel_bina:,.4f} ₺")
m_col3.metric("Güncel Pafta B.F.", f"{guncel_pafta:,.4f} ₺", delta=mesh_durumu.split("(")[0].strip(), delta_color="normal")
st.markdown("---")

# --- DEVASA SENARYO SEÇİM ALANI ---
st.markdown("<h2 style='text-align: center; color: #ff4b4b; font-size: 36px;'>🛠️ İŞ TÜRÜNÜ SEÇİNİZ 🛠️</h2>", unsafe_allow_html=True)
is_senaryosu = st.radio(
    "Senaryo Seçimi",
    ("🔴 İYİLEŞTİRME İŞİ (Kontrol ve Revizyon)", "🟢 OLUŞTURMA İŞİ (Sıfırdan Üretim)"),
    horizontal=True,
    label_visibility="collapsed"
)
st.markdown("---")

# --- ÇOKLU İŞ GİRİŞİ (MAX 5) ---
st.subheader("📋 İş Listesi ve Miktarlar (Maksimum 5 İş)")
if "İYİLEŞTİRME" in is_senaryosu:
    # İyileştirme için sütunlar
    init_data = pd.DataFrame([{"İş Adı": "", "Pafta Sayısı": 0, "Toplam Kontrol Bina": 0, "Sıfırdan Üretilecek Bina": 0}] * 5)
else:
    # Oluşturma için sütunlar
    init_data = pd.DataFrame([{"İş Adı": "", "Pafta Sayısı": 0, "Toplam Bina": 0}] * 5)

edited_df = st.data_editor(init_data, use_container_width=True, hide_index=True)

# --- HESAPLAMA FONKSİYONU ---
def maliyet_hesapla(row, senaryo):
    pafta = float(row.get("Pafta Sayısı", 0))
    is_ciplak = 0.0
    
    if "İYİLEŞTİRME" in senaryo:
        top_b = float(row.get("Toplam Kontrol Bina", 0))
        sif_b = float(row.get("Sıfırdan Üretilecek Bina", 0))
        esk_b = top_b - sif_b
        
        # Pafta kalemleri (%85 ağırlık, kalem 4 yarı yarıya düşüyor)
        is_ciplak += excel_yuvarla(excel_yuvarla(guncel_pafta * 0.85, 2) * pafta, 2)
        # Bina kalemleri
        is_ciplak += excel_yuvarla(excel_yuvarla(guncel_bina * 0.60, 2) * top_b, 2) # 1, 8, 9
        is_ciplak += excel_yuvarla(excel_yuvarla(guncel_bina * 0.40, 2) * sif_b, 2) # 6
        is_ciplak += excel_yuvarla(excel_yuvarla(guncel_bina * 0.20, 3) * esk_b, 2) # 7 (Extra %20)
    else:
        top_b = float(row.get("Toplam Bina", 0))
        is_ciplak += excel_yuvarla(excel_yuvarla(guncel_pafta, 2) * pafta, 2)
        is_ciplak += excel_yuvarla(excel_yuvarla(guncel_bina, 2) * top_b, 2)
        
    kar = excel_yuvarla(is_ciplak * (kar_orani / 100), 2)
    toplam = is_ciplak + kar
    return is_ciplak, kar, toplam

# --- EXCEL ŞABLONUNA AKTARMA VE İNDİRME ---
if st.button("🚀 Hesapla ve Excel Şablonuna Aktar (sablon_ym.xlsx)"):
    try:
        # Mevcut şablonu yükle
        wb = openpyxl.load_workbook("sablon_ym.xlsx")
        ws = wb["ÇIKTI"]
        
        hesap_sayisi = 0
        for i, row in edited_df.iterrows():
            if not row["İş Adı"]: continue # İş adı boşsa atla
            
            hesap_sayisi += 1
            idx = i + 5 # 5. satırdan başlar (B5, B6...)
            
            ham_maliyet, kar, toplam_maliyet = maliyet_hesapla(row, is_senaryosu)
            
            # Belirlediğiniz hücrelere yazım
            ws[f"B{idx}"] = row["İş Adı"]
            ws[f"F{idx}"] = row["Pafta Sayısı"]
            ws[f"J{idx}"] = ham_maliyet
            ws[f"K{idx}"] = kar
            ws[f"L{idx}"] = toplam_maliyet
            
        # Bellekte dosyayı oluştur
        buffer = io.BytesIO()
        wb.save(buffer)
        
        st.success(f"✅ {hesap_sayisi} adet iş başarıyla hesaplandı ve şablona işlendi!")
        st.download_button(
            label="📥 Hazır Excel Şablonunu İndir",
            data=buffer.getvalue(),
            file_name=f"AYAP_Maliyet_Ciktisi_{secilen_donem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
    except FileNotFoundError:
        st.error("❌ 'sablon_ym.xlsx' dosyası ana dizinde bulunamadı. Lütfen şablon dosyasını yükleyin.")
    except Exception as e:
        st.error(f"❌ Bir hata oluştu: {e}")

st.markdown("---")
st.info("💡 Not: Bu araç girdiğiniz verileri kuruşu kuruşuna Excel'in yuvarlama mantığıyla hesaplar.")
